# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import AccessError, UserError, ValidationError
from datetime import date, timedelta
import xlsxwriter
import io
import base64
import openpyxl
from pathlib import Path
import xlrd
from xlrd import open_workbook


class HotelRoom(models.Model):
    _name = 'hotel.room'
    _rec_name = 'room_type_id'
    _rec_name = 'room_no'

    room_no = fields.Char('Room Number', readonly=True, required=True, index=True, default='New')
    room_type_id = fields.Many2one('room.type', string='Room Type Id')
    floor = fields.Selection([('one', '1'), ('two', '2'), ('three', '3')], string='Floor')
    room_size = fields.Integer('Room Size')
    inquiry_id = fields.Many2one('hotel.inquiry', 'Inquiry')
    book_room = fields.Boolean('Room Book')
    price = fields.Integer('Price')
    room_state = fields.Selection([
        ('draft', 'Draft'),
        ('allocated', 'Allocated')
    ], default='draft')

    def allocated_progressbar(self):
        for rec in self:
            rec.room_state = 'allocated'

    @api.model
    def create(self, vals):
        if vals.get('room_no', 'New') == 'New':
            vals['room_no'] = self.env['ir.sequence'].next_by_code('hotel.room') or 'New'
        result = super(HotelRoom, self).create(vals)
        return result

    def get_registration(self):
        print("----->\n\n\n\n\n")
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Registration',
            'view_type': 'form',
            'view_mode': 'tree',
            'res_model': 'hotel.registration',
            'domain': [('line_ids.room_id', '=', self.id)],
        }

    def get_inquiry(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': 'Inquiry',
            'view_type': 'form',
            'view_mode': 'tree',
            'res_model': 'hotel.inquiry',
            # 'domain': [('line_ids.room_id', '=', self.id)],
        }


class HotelRegistration(models.Model):
    _name = 'hotel.registration'
    _rec_name = 'name'

    name = fields.Char('Name')
    register_no = fields.Char('Register Number', readonly=True, required=True, index=True, default='New')
    phone = fields.Integer('Contact Number')
    dob = fields.Date('Birth Date')
    doc_ids = fields.One2many('register.document', 'registration_id', string='Documents', required=True)
    date_from = fields.Date('Start Date')
    date_to = fields.Date('End Date')
    total = fields.Float('Total')
    state = fields.Selection([
        ('draft', 'Draft'),
        ('process', 'Process'),
        ('done', 'Done'),
        ('cancel', 'Cancel')
    ], default='draft')
    line_ids = fields.One2many('room.guest.line', 'reg_id', string='Room Customer Guest')

    def done_progressbar(self):
        for rec in self:
            rec.state = 'done'
            rec.line_ids.room_id.room_state = 'allocated'

    def process_progressbar(self):
        for rec in self:
            rec.state = 'process'

    def cancel_progressbar(self):
        for rec in self:
            rec.state = 'cancel'

    def action_registration_send(self):
        ''' Opens a wizard to compose an email, with relevant mail template loaded by default '''
        self.ensure_one()
        ir_model_data = self.env['ir.model.data']

        try:
            template_id = self.env['ir.model.data'].xmlid_to_res_id('hotel_management.email_template_reg',
                                                                    raise_if_not_found=False)
            print(template_id, "template_id\n\n\n\n\n")
        except ValueError:
            print("\n\n\n\n value error \n\n\n\n\n\n")
            template_id = False
        try:
            compose_form_id = ir_model_data.get_object_reference('mail', 'email_compose_message_wizard_form')[1]
        except ValueError:
            compose_form_id = False

        ctx = {
            'default_model': 'hotel.registration',
            'default_res_id': self.ids[0],
            'default_use_template': bool(template_id),
            'default_template_id': template_id,
            'default_composition_mode': 'comment',
        }
        return {
            'name': _('Compose Email'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'mail.compose.message',
            'views': [(compose_form_id, 'form')],
            'view_id': compose_form_id,
            'target': 'new',
            'context': ctx,
        }

    @api.model
    def create(self, vals):
        if vals.get('register_no', 'New') == 'New':
            vals['register_no'] = self.env['ir.sequence'].next_by_code('hotel.registration') or 'New'
        result = super(HotelRegistration, self).create(vals)
        return result

    @api.model
    def _registration_cancel(self):
        print("--\n\n\n--->", self)
        var = self.env['hotel.registration'].search([('state', '!=', 'done')])
        for i in var:
            if ((date.today()) - (i.create_date)) > timedelta(days=3):
                i.state = 'cancel'

                print("\n\n\n\n\n\n\n\n--")
                print("\n\n\n\n\n\n\n\n--")
            else:
                print("\n\n\n\n\n\n\n\n--")


class RegisterDocument(models.Model):
    _name = 'register.document'

    name = fields.Char('Name')
    date = fields.Date('Date')
    doc = fields.Binary('Document', required=True)
    registration_id = fields.Many2one('hotel.registration', string='Registration Id')


class RoomType(models.Model):
    _name = 'room.type'
    _rec_name = 'room_type'

    room_type = fields.Char('Room Type')


class HotelInquiry(models.Model):
    _name = 'hotel.inquiry'
    _description = 'Inquiry about Hotel and Rooms Avaibility'
    _rec_name = 'customer'

    customer = fields.Many2one('res.partner', string='Customer', required=True)
    start_date = fields.Date('Start Date')
    end_date = fields.Date('End Date')
    room_types = fields.Many2one('room.type', 'Room Type')
    room_size_id = fields.Integer('Room Size')
    room_ids = fields.One2many('hotel.room', 'inquiry_id', string='Room')
    total_price = fields.Integer('Total Price', compute='compute_price')

    @api.onchange('room_ids')
    def compute_price(self):
        total = 0
        for record in self.room_ids:
            if record.book_room:
                total += record.price
        self.write({
            'total_price': total
        })

    def search_room_available(self):
        check_room = self.env['hotel.room'].search(
            [('room_state', '=', 'draft'), ('room_type_id', '=', self.room_types.id),
             ('room_size', '>=', self.room_size_id)])
        print("\n\n\n\n\n---->", check_room)
        if check_room:
            self.room_ids = [(6, 0, [])]
            print("-----t-->", self.room_ids)
            self.write({'room_ids': check_room})
            return


        else:
            raise ValidationError(_('No room Available'))
            print("__________----------________________-------------________________----------_______________--")

    def get_record(self):
        records = self.env['hotel.inquiry'].browse('active_ids')
        print("---------->", records)

        res = []
        for i in self:
            for j in i.room_ids:
                if j.book_room:
                    res.append({'room_id': j.id})
                    print("----------->", res)

        contexts = {

            'default_line_ids': res,
            'default_name': self.customer.name,
            'default_date_from': self.start_date,
            'default_date_to': self.end_date,
            'default_total': self.total_price
        }

        if records:
            return {
                'res_model': 'hotel.registration',
                'type': 'ir.actions.act_window',
                'context': contexts,
                'view_mode': 'form',
                'view_type': 'form',
                'view_id': self.env.ref("hotel_management.registration_form_view").id,
                'target': 'new'
            }


class RoomGuestLine(models.Model):
    _name = 'room.guest.line'
    _description = 'Inquiry about customer and their guests Avaibility'
    _rec_name = 'room_id'

    room_id = fields.Many2one("hotel.room", 'Room')
    guest_ids = fields.Many2many('res.partner', string='Guests')
    reg_id = fields.Many2one('hotel.registration', 'Register Id')


class Registration(models.AbstractModel):
    _name = 'report.hotel_management.report_registration'
    _description = 'get values data'

    def _get_method(self, doc):
        print("---\\n\n\n\n\n---")
        return "5"

    @api.model
    def _get_report_values(self, docids, data=None):
        print("\n\n\n\n\n\n\n\n\n re")
        docs = self.env['hotel.registration'].browse(docids)
        return {
            'doc_ids': docids,
            'doc_model': 'hotel.registration',
            'docs': docs,
            'data': data,
            'get_method': self._get_method,
        }


class RegistrationReport(models.TransientModel):
    _name = 'registration.report'
    _description = 'Registration Xls Report'

    start_date = fields.Date('Start Date')
    end_date = fields.Date('End Date')

    def generate_xlsx_report(self):
        print("------------>", self)
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet()
        cell_format = workbook.add_format()
        number_format = workbook.add_format({'num_format': 'dd/mm/yyyy'})
        cell_format.set_align('center')
        cell_format.set_align('vcenter')
        cell_format = workbook.add_format({'bold': True, 'font_color': 'black', 'bg_color': 'red'})
        cell_format.set_italic()
        cell_format.set_underline()
        worksheet.write(0, 1, "Start Date", cell_format)
        worksheet.write(0, 2, "End Date", cell_format)
        worksheet.write(0, 3, "Register Number", cell_format)
        worksheet.write(0, 4, "Customer", cell_format)
        worksheet.write(0, 5, "Date of Birth", cell_format)
        worksheet.write(0, 6, "Total", cell_format)

        worksheet.set_row(0, 20)
        worksheet.set_column('B:K', 20)

        row = 1
        col = 1

        # Iterate over the data and write it out row by row.
        records = self.env['hotel.registration'].search([
            ('date_from', '>=', self.start_date),
            ('date_to', '<=', self.end_date)
        ])
        print("-------records------>", records)

        for rec in records:
            worksheet.write(row, col, rec.date_from, number_format)
            print("-----start date----->", rec.date_from)
            worksheet.write(row, col + 1, rec.date_to, number_format)
            worksheet.write(row, col + 2, rec.register_no)
            worksheet.write(row, col + 3, rec.name)
            worksheet.write(row, col + 4, rec.dob, number_format)
            worksheet.write(row, col + 5, rec.total, number_format)

            row += 1
        workbook.close()

        output.seek(0)
        attch = self.env['ir.attachment'].create(
            {'name': 'Registrations.xlsx', 'datas': base64.b64encode(output.read())})
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/' + str(attch.id) + '?download=True',
            'target': self
        }


class ImportRoom(models.TransientModel):
    _name = 'import.room'

    file = fields.Binary(string='Upload File')

    def get_room_data(self):
        print(self.file)
        nm = xlrd.open_workbook(file_contents=base64.decodebytes(self.file))
        for sheet in nm.sheets():
            for row in range(sheet.nrows):
                if row != 0:
                    room_type_id = sheet.cell(row, 0).value
                    floor = sheet.cell(row, 1).value
                    room_size = sheet.cell(row, 2).value
                    room_state = sheet.cell(row, 3).value
                    price = sheet.cell(row, 4).value

                    self.env['hotel.room'].create({

                        'room_no': 'New',
                        'room_type_id': int(room_type_id),
                        'floor': str(floor),
                        'room_size': int(room_size),
                        'room_state': str(room_state),
                        'price': int(price)
                    })
                else:
                    print("---\n\n\n\n\\n\n\n\n---")


class SaleOrder(models.Model):
    _inherit = 'sale.order'
    _description = 'Inherit sale order line and add button above order lines'

    def get_line_wizard(self):
        print("-------\n\n\n-->", self)
        return {
            'res_model': 'import.line',
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'view_type': 'form',
            'view_id': self.env.ref("hotel_management.import_line_form_view").id,
            'target': 'new'
        }


class ImportLine(models.TransientModel):
    _name = 'import.line'

    document = fields.Binary(string='Upload File')

    def get_import_line(self):
        print(self.document)
        ab = xlrd.open_workbook(file_contents=base64.decodebytes(self.document))
        for sheet in ab.sheets():
            for row in range(1, sheet.nrows):
                product_name = sheet.cell(row, 0).value
                desc = sheet.cell(row, 1).value
                active = self.env.context.get('active_id')
                exist = self.env['product.product'].search([(
                    'name', '=', product_name
                )], limit=1)
                if len(exist.ids) >= 1:
                    product = exist.id
                else:
                    product = self.env['product.product'].create({
                        'name': str(product_name)
                    }).id

                self.env['sale.order.line'].create({
                    "order_id": active,
                    "product_id": product,
                    "name": desc
                })
